from apps.home import blueprint
from apps import db
from apps.authentication.forms import CreateNodeDataForm
from apps.authentication.models import Node1,Users,Node2,CPU,WiFi,Memory
from flask import render_template, request,Response
from flask_login import login_required
from paho.mqtt import client as mqtt_client
from jinja2 import TemplateNotFound
import random
import datetime
import json
import sqlite3
import io
import csv
import datetime
from openpyxl import Workbook
broker = 'arzethamescarloswsnproyecto2.com'
port = 1883
topic = "values"
# generate client ID with pub prefix randomly
client_id = f'python-mqtt-{random.randint(0, 1000)}'
username = 'emqx'
password = '**********'
flag=0
def connect_mqtt():
    def on_connect(client, userdata, flags, rc):
        if rc == 0:
            print("Connected to MQTT Broker!")
        else:
            print("Failed to connect, return code %d\n", rc)
   
    client = mqtt_client.Client(client_id)
    # client.tls_set(ca_certs='./server-ca.crt')
    client.username_pw_set(username, password)
    client.on_connect = on_connect
    client.connect(broker, port)
    return client
def subscribe(client: mqtt_client):
    def on_message(client, userdata, msg):
        print(f"Received `{msg.payload.decode()}` from `{msg.topic}` topic")
        x=msg.payload.decode()
        '''float(y["nodeId"]),float(y["rssi"]),float(y["snr"]),float(y["data"])'''
        y=json.loads(x)
        a=[int(y["usecpu0"]),int(y["usecpu1"]),str(y["model"]),int(y["cores"]),int(y["features"])
           ,int(y["revision"]),int(y["clkcpu"]),int(y["xtalclk"]),float(y["value1"]),float(y["value2"])
           ,float(y["value3"]),float(y["value4"]),float(y["value5"]),float(y["value6"]),float(y["valuea1"]),float(y["valuea2"])
           ,float(y["valuea3"]),float(y["valuea4"]),float(y["valuea5"]),float(y["valuea6"]),int(y["rssiLoRa"]),float(y["snrLoRa"])
           ,int(y["rssiLoRa2"]),float(y["snrLoRa2"]),str(y["ssid"]),str(y["scanmethod"]),str(y["channel"]),str(y["mac"]),str(y["bssid"]),str(y["sortmethod"]),int(y["rssithr"]),str(y["authmode"])
           ,str(y["protocol"]),str(y["bandwidth"]),int(y["rssireal"])]
        fecha_actual = datetime.datetime.now()
        con = sqlite3.connect("./apps/db.sqlite3")
        cur = con.cursor()
        cur.execute(f"INSERT INTO Node2(nodeId, gatewayId, rssi, snr, data1, data2, time) VALUES (246, 254, ?, ?, ?, ?, ?)", (a[22], a[23], a[14], a[15], fecha_actual))
        con.commit()
        cur.execute(f"INSERT INTO Node1(nodeId,gatewayId,rssi,snr,data1,data2,data3,data4,data5,data6,time) VALUES (245,254,?,?,?,?,?,?,?,?,?)",(a[20],a[21],a[8],a[9],a[10],a[11],a[12],a[13],fecha_actual))
        con.commit()
        cur.execute(f"INSERT INTO CPU(procpu,appcpu,model,cores,features,revision,clkcpu,xtalcpu,time) VALUES (?,?,?,?,?,?,?,?,?)",(a[0],a[1],a[2],a[3],a[4],a[5],a[6],a[7],fecha_actual))
        con.commit()
        cur.execute(f"INSERT INTO WiFi(ssid,scanmethod,channel,mac,bssid,sortmethod,rssithr,authmode,protocol,bandwidth,rssireal,time) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",(a[24],a[25],a[26],a[27],a[28],a[29],a[30],a[31],a[32],a[33],a[34],fecha_actual))
        con.commit()
        cur.execute("INSERT INTO Memory(dramdata, drambss, iramtext, iramvectors, flashtext, flashsize, time) VALUES (?, ?, ?, ?, ?, ?, ?)", ('15588', '16528', '89471', '1027', '656183', '167742', fecha_actual))
        con.commit()
        con.close()
    client.subscribe(topic)
    client.on_message = on_message

@blueprint.route('/index')
@login_required
def index():
    global flag
    if(flag==0):
        client = connect_mqtt()
        subscribe(client)
        client.loop_start()
        flag=1
    '''dictNode={}
    dictNode['nodeId']=2
    dictNode['rssi']=3
    dictNode['snr']=14
    dictNode['data']=3.14
    print(dictNode)
    node = Node(**dictNode)
    db.session.add(node)
    db.session.commit()'''
    '''b=Node.query.order_by(Node.id).all()
    rssiValues=[]
    for i in range(len(b)):
        rssiValues.append(b[i].rssi)
    print(rssiValues)'''
    return render_template('home/index.html', segment='index')

@blueprint.route('/download/report/csv1')
@login_required
def download_report_csv():
    output = io.StringIO()
    writer = csv.writer(output)
    line = ['nodeId','gatewayId','rssi','snr','CO2','TVOC','uvLevel','refLevel','outputVoltage','uvIntensity','time']
    b=Node1.query.order_by(Node1.id).all()
    writer.writerow(line)
    for i in range(len(b)):
        line = [str(b[i].nodeId) + ',' + str(b[i].gatewayId) + ',' + str(b[i].rssi) + ',' + str(b[i].snr) + ',' + str(b[i].data1)+',' + str(b[i].data2)+',' + str(b[i].data3)+',' + str(b[i].data4)+',' + str(b[i].data5)+',' + str(b[i].data6)+',' + str(b[i].time)]
        writer.writerow(line)
    output.seek(0)
    return Response(output, mimetype="text/csv", headers={"Content-Disposition":"attachment;filename=nodeUVCCS811IoT.csv"})
    
@blueprint.route('/download/report/xlsx1')
@login_required
def download_report_xlsx():
    b=Node1.query.order_by(Node1.id).all()
    wb = Workbook()
    ws = wb.active
    ws.append( ['nodeId','gatewayId','rssi','snr','CO2','TVOC','uvLevel','refLevel','outputVoltage','uvIntensity','time'])
    for i in range(len(b)):
        line = [str(b[i].nodeId) + ',' + str(b[i].gatewayId) + ',' + str(b[i].rssi) + ',' + str(b[i].snr) + ',' + str(b[i].data1)+',' + str(b[i].data2)+',' + str(b[i].data3)+',' + str(b[i].data4)+',' + str(b[i].data5)+',' + str(b[i].data6)+',' + str(b[i].time)]
        ws.append(line)
    memoria = io.BytesIO()
    wb.save(memoria)
    memoria.seek(0)
    return Response(
        memoria,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=nodeUVCCS811IoT.xlsx"}
    ) 

@blueprint.route('/download/report/csv2')
@login_required
def download_report_csv2():
    output = io.StringIO()
    writer = csv.writer(output)
    line = ['nodeId','gatewayId','rssi','snr','CO','NO2','time']
    b=Node2.query.order_by(Node2.id).all()
    writer.writerow(line)
    for i in range(len(b)):
        line = [str(b[i].nodeId) + ',' + str(b[i].gatewayId) + ',' + str(b[i].rssi) + ',' + str(b[i].snr) + ',' + str(b[i].data1)+',' + str(b[i].data2)+',' + str(b[i].time)]
        writer.writerow(line)
    output.seek(0)
    return Response(output, mimetype="text/csv", headers={"Content-Disposition":"attachment;filename=nodeMiCS6814IoT.csv"})
    
@blueprint.route('/download/report/xlsx2')
@login_required
def download_report_xlsx2():
    b=Node2.query.order_by(Node2.id).all()
    wb = Workbook()
    ws = wb.active
    ws.append( ['nodeId','gatewayId','rssi','snr','CO','NO2','time'])
    for i in range(len(b)):
        line = [str(b[i].nodeId) + ',' + str(b[i].gatewayId) + ',' + str(b[i].rssi) + ',' + str(b[i].snr) + ',' + str(b[i].data1)+',' + str(b[i].data2)+',' + str(b[i].time)]
        ws.append(line)
    memoria = io.BytesIO()
    wb.save(memoria)
    memoria.seek(0)
    return Response(
        memoria,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=nodeMiCS6814IoT.xlsx"}
    ) 
@blueprint.route('/download/report/csv3')
@login_required
def download_report_csv3():
    output = io.StringIO()
    writer = csv.writer(output)
    line = ['Procpu','Appcpu','Model','Cores','Features','Revision','CLK CPU','XTAL CPU']
    b=CPU.query.order_by(CPU.id).all()
    writer.writerow(line)
    for i in range(len(b)):
        line = [str(b[i].procpu) + ',' + str(b[i].appcpu) + ',' + str(b[i].model) + ',' + str(b[i].cores) + ',' + str(b[i].features)+',' + str(b[i].revision)+',' + str(b[i].clkcpu)+',' + str(b[i].xtalcpu)]
        writer.writerow(line)
    output.seek(0)
    return Response(output, mimetype="text/csv", headers={"Content-Disposition":"attachment;filename=gatewayCPU.csv"})
    
@blueprint.route('/download/report/xlsx3')
@login_required
def download_report_xlsx3():
    b=CPU.query.order_by(CPU.id).all()
    wb = Workbook()
    ws = wb.active
    ws.append(['Procpu','Appcpu','Model','Cores','Features','Revision','CLK CPU','XTAL CPU'])
    for i in range(len(b)):
        line = [str(b[i].procpu) + ',' + str(b[i].appcpu) + ',' + str(b[i].model) + ',' + str(b[i].cores) + ',' + str(b[i].features)+',' + str(b[i].revision)+',' + str(b[i].clkcpu)+',' + str(b[i].xtalcpu)]
        ws.append(line)
    memoria = io.BytesIO()
    wb.save(memoria)
    memoria.seek(0)
    return Response(
        memoria,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=gatewayCPU.xlsx"}
    ) 
@blueprint.route('/download/report/csv4')
@login_required
def download_report_csv4():
    output = io.StringIO()
    writer = csv.writer(output)
    line = ['SSID','SCAN Method','Channel','MAC','BSSID','Sort Method','RSSI Threshold','Auth mode','Protocol','RSII Real','Time']
    b=WiFi.query.order_by(WiFi.id).all()
    writer.writerow(line)
    for i in range(len(b)):
        line = [str(b[i].ssid) + ',' + str(b[i].scanmethod) + ',' + str(b[i].channel) + ',' + str(b[i].mac) + ',' + str(b[i].bssid)+',' + str(b[i].sortmethod)+',' + str(b[i].rssithr)+',' + str(b[i].authmode)+','+str(b[i].protocol)+','+str(b[i].bandwidth)+','+str(b[i].rssireal)+','+str(b[i].time)]
        writer.writerow(line)
    output.seek(0)
    return Response(output, mimetype="text/csv", headers={"Content-Disposition":"attachment;filename=gatewayWiFi.csv"})
    
@blueprint.route('/download/report/xlsx4')
@login_required
def download_report_xlsx4():
    b=WiFi.query.order_by(WiFi.id).all()
    wb = Workbook()
    ws = wb.active
    ws.append(['SSID','SCAN Method','Channel','MAC','BSSID','Sort Method','RSSI Threshold','Auth mode','Protocol','RSII Real','Time'])
    for i in range(len(b)):
        line = [str(b[i].ssid) + ',' + str(b[i].scanmethod) + ',' + str(b[i].channel) + ',' + str(b[i].mac) + ',' + str(b[i].bssid)+',' + str(b[i].sortmethod)+',' + str(b[i].rssithr)+',' + str(b[i].authmode)+','+str(b[i].protocol)+','+str(b[i].bandwidth)+','+str(b[i].rssireal)+','+str(b[i].time)]
        ws.append(line)
    memoria = io.BytesIO()
    wb.save(memoria)
    memoria.seek(0)
    return Response(
        memoria,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=gatewayWiFi.xlsx"}
    ) 		
@blueprint.route('/download/report/csv5')
@login_required
def download_report_csv5():
    output = io.StringIO()
    writer = csv.writer(output)
    line = ['DRAM data','DRAM bss','IRAM text','IRAM vectors','Flash text','Flash size','Time']
    b=Memory.query.order_by(Memory.id).all()
    writer.writerow(line)
    for i in range(len(b)):
        line = [str(b[i].dramdata) + ',' + str(b[i].drambss) + ',' + str(b[i].iramtext) + ',' + str(b[i].iramvectors) + ',' + str(b[i].flashtext)+',' + str(b[i].flashsize)+','+str(b[i].time)]
        writer.writerow(line)
    output.seek(0)
    return Response(output, mimetype="text/csv", headers={"Content-Disposition":"attachment;filename=gatewayMemory.csv"})
    
@blueprint.route('/download/report/xlsx5')
@login_required
def download_report_xlsx5():
    b=Memory.query.order_by(Memory.id).all()
    wb = Workbook()
    ws = wb.active
    ws.append(['DRAM data','DRAM bss','IRAM text','IRAM vectors','Flash text','Flash size','Time'])
    for i in range(len(b)):
        line = [str(b[i].dramdata) + ',' + str(b[i].drambss) + ',' + str(b[i].iramtext) + ',' + str(b[i].iramvectors) + ',' + str(b[i].flashtext)+',' + str(b[i].flashsize)+','+str(b[i].time)]
        ws.append(line)
    memoria = io.BytesIO()
    wb.save(memoria)
    memoria.seek(0)
    return Response(
        memoria,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=gatewayMemory.xlsx"}
    ) 	

 
@blueprint.route('/<template>')
@login_required
def route_template(template):
    try:

        if not template.endswith('.html'):
            template += '.html'
        segment = get_segment(request)
        if template=="node1metrics.html":
            node1=Node1.query.order_by(Node1.id).all()
            return render_template("home/" + template, segment=segment,node1=node1)
        if template=="node2metrics.html":
            node2=Node2.query.order_by(Node2.id).all()
            return render_template("home/" + template, segment=segment,node2=node2)
        if template=="ui-gatewaymetrics.html":
            cpu=CPU.query.order_by(CPU.id).all()
            wifi=WiFi.query.order_by(WiFi.id).all()
            memory=Memory.query.order_by(Memory.id).all()
            return render_template("home/" + template, segment=segment,gateway=cpu,wifi=wifi,memory=memory)
            
        return render_template("home/" + template, segment=segment)

    except TemplateNotFound:
        return render_template('home/page-404.html'), 404

    except:
        return render_template('home/page-500.html'), 500
def get_segment(request):
    try:
        segment = request.path.split('/')[-1]

        if segment == '':
            segment = 'index'

        return segment

    except:
        return None
